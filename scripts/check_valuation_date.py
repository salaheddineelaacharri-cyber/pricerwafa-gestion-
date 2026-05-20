"""Inspecte la date de valorisation et les prix dans le classeur de référence."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

wb = load_workbook(wb_path, data_only=True, keep_vba=True)
print("Sheets:", wb.sheetnames)

# Look on each sheet for cells with valuation date
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            sv = str(v).lower()
            if "valo" in sv or "date de val" in sv or "valuation" in sv or "Date" == str(v):
                # Print this cell + neighbour
                print(f"  [{sn}] {cell.coordinate} = {cell.value!r}  (next-> {ws.cell(cell.row, cell.column+1).value!r})")
