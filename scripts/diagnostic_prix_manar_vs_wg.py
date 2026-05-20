from __future__ import annotations

import json
import math
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
import pandas as pd

from backend import main as api


def _norm_code(v: Any) -> str:
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _excel_prices() -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    wb_v = openpyxl.load_workbook(ROOT / "2026-PRICER_WG_CORRIGE .xlsm", read_only=True, keep_vba=True, data_only=True)
    wb_f = openpyxl.load_workbook(ROOT / "2026-PRICER_WG_CORRIGE .xlsm", read_only=True, keep_vba=True, data_only=False)
    try:
        ws_v = wb_v.worksheets[1]
        ws_f = wb_f.worksheets[1]
        for row in ws_v.iter_rows():
            code = _norm_code(row[0].value)
            if not code:
                continue
            price = row[1].value
            if not isinstance(price, (int, float)) or not math.isfinite(float(price)):
                continue
            r = row[0].row
            out[code] = {
                "prix_wg": round(float(price), 2),
                "prix_wg_raw": float(price),
                "row": r,
                "formule": str(ws_f.cell(r, 24).value or ""),
                "rendement": ws_v.cell(r, 4).value,
                "base": ws_v.cell(r, 25).value,
                "mode": ws_v.cell(r, 22).value,
            }
    finally:
        wb_v.close()
        wb_f.close()
    return out


def main() -> None:
    prix = pd.read_excel(ROOT / "prix manarrr.xlsx")
    prix.columns = [str(c).strip().lower() for c in prix.columns]
    prix_rows = [
        {"code": _norm_code(r["titre"]), "manar": round(float(r["valo"]), 2)}
        for _, r in prix.iterrows()
        if str(r.get("titre") or "").strip()
    ]
    excel_map = _excel_prices()

    pillars = api._extraire_piliers_depuis_histo(ROOT, "2026-03-26", "MAR_JJ")
    curve = api.CurveRequest(
        short=[api.PillarShort(**p) for p in pillars["short"]],
        long=[api.PillarLong(**p) for p in pillars["long"]],
        joint_days=float(pillars.get("joint_days", 325.0)),
        max_days=11000,
        step_short=50,
        step_long=100,
    )
    req = api.MarcheValorizeRequest(
        valuation_date="2026-03-26",
        curve=curve,
        prix_manarr_pricer_tous=True,
    )
    res = api.marche_valorize(req)
    if hasattr(res, "body"):
        res = json.loads(res.body)
    moteur = {
        _norm_code(r.get("titre")): r
        for r in res.get("prix_manarr", [])
        if _norm_code(r.get("titre"))
    }

    rows = []
    for item in prix_rows:
        code = item["code"]
        manar = item["manar"]
        wg = excel_map.get(code)
        mot = moteur.get(code, {})
        prix_moteur = mot.get("prix_arrondi")
        profil = mot.get("profil_metier")
        ecart_moteur = None if prix_moteur is None else round(float(prix_moteur) - manar, 2)
        ecart_wg = None if not wg else round(float(wg["prix_wg"]) - manar, 2)
        rows.append(
            {
                "code": code,
                "profil": profil,
                "manar": manar,
                "moteur": prix_moteur,
                "ecart_moteur": ecart_moteur,
                "wg": None if not wg else wg["prix_wg"],
                "ecart_wg": ecart_wg,
                "formule": None if not wg else wg["formule"],
            }
        )
    bad = [r for r in rows if r["ecart_moteur"] is None or abs(float(r["ecart_moteur"])) > 0.02]
    print(f"TOTAL={len(rows)} BAD_MOTEUR={len(bad)}")
    by_profile: dict[str, int] = {}
    for r in bad:
        by_profile[str(r.get("profil") or "profil inconnu")] = by_profile.get(str(r.get("profil") or "profil inconnu"), 0) + 1
    for profil, n in sorted(by_profile.items(), key=lambda kv: (-kv[1], kv[0]))[:30]:
        print(f"{n:3d} {profil}")
    print("DETAILS_TOP")
    for r in bad[:120]:
        print(r)


if __name__ == "__main__":
    main()
