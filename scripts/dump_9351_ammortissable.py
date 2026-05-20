"""Dump le bloc Ammortissable pour 9351 — formules + valeurs calculées."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

wb_v = load_workbook(wb_path, data_only=True, keep_vba=True)
wb_f = load_workbook(wb_path, data_only=False, keep_vba=True)

ws_v = wb_v["Ammortissable"]
ws_f = wb_f["Ammortissable"]

# 9351 — search code in column B/C
target_rows: list[int] = []
for row in ws_v.iter_rows(min_col=1, max_col=5, values_only=False):
    for cell in row:
        if cell.value is not None and str(cell.value).strip() == "9351":
            target_rows.append(cell.row)

if not target_rows:
    print("9351 introuvable")
else:
    r0 = min(target_rows)
    last_col = min(ws_v.max_column, 35)
    for r in range(r0, r0 + 16):
        for c in range(1, last_col + 1):
            f = ws_f.cell(r, c).value
            v = ws_v.cell(r, c).value
            if f is None and v is None:
                continue
            col_letter = chr(ord('A') + c - 1) if c <= 26 else 'A' + chr(ord('A') + c - 27)
            is_formula = isinstance(f, str) and f.startswith("=")
            if is_formula:
                print(f"  {col_letter}{r}: F={f!r}  V={v!r}")
            else:
                print(f"  {col_letter}{r}: V={v!r}")
        print()
