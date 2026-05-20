"""Dump des **formules** de chaque cellule de l'Ammortissable 9500 — pour comprendre Excel."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

wb_f = load_workbook(wb_path, data_only=False, keep_vba=True)
wb_v = load_workbook(wb_path, data_only=True, keep_vba=True)
ws_f = wb_f["Ammortissable"]
ws_v = wb_v["Ammortissable"]

# Code 9500 starts at row 1077; we look at 1080..1092
for r in range(1077, 1093):
    print(f"\n--- Row {r} ---")
    for c in range(1, 35):  # A..AI
        f = ws_f.cell(r, c).value
        v = ws_v.cell(r, c).value
        if f is None and v is None:
            continue
        col = chr(ord('A') + c - 1) if c <= 26 else 'A' + chr(ord('A') + c - 27)
        is_formula = isinstance(f, str) and f.startswith("=")
        if is_formula:
            print(f"  {col}{r}: FORMULA={f!r}  VALUE={v!r}")
        else:
            print(f"  {col}{r}: VALUE={v!r}")
