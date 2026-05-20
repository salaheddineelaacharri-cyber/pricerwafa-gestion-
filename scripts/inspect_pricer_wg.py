"""Outil d'analyse du classeur 2026-PRICER_WG_CORRIGE pour les codes 9424/9500.

Cherche les blocs 'titre' dans la feuille Ammortissable et imprime, pour chacun,
le contenu des cellules autour (formules + valeurs calculées via data_only).
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WB_PATH = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

TARGETS = {9424, 9500, 9351, 9651, 101006}


def find_in_sheet(ws) -> list[tuple[str, int, int, object]]:
    """Retourne les cellules contenant un code cible."""
    found: list[tuple[str, int, int, object]] = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            try:
                f = float(v) if isinstance(v, (int, float)) else float(str(v).strip())
            except Exception:
                continue
            iv = int(f)
            if abs(f - iv) < 1e-9 and iv in TARGETS:
                found.append((ws.title, cell.row, cell.column, v))
    return found


def main() -> None:
    wb_f = load_workbook(WB_PATH, data_only=False, read_only=True)
    wb_v = load_workbook(WB_PATH, data_only=True, read_only=True)
    print("Sheets:", wb_f.sheetnames)
    candidate_sheets = [
        "Ammortissable",
        "ONCF",
        "CIH",
        "Echéanciers des Parts",
        "Echeancier de parts 2",
        "FEC9149-2140",
        "ammc",
    ]
    for sn in candidate_sheets:
        if sn not in wb_f.sheetnames:
            continue
        ws = wb_f[sn]
        hits = find_in_sheet(ws)
        if hits:
            print(f"\n=== {sn} : {len(hits)} occurrences ===")
            for s, r, c, v in hits:
                print(f"  {s}!{ws.cell(row=r, column=c).coordinate} = {v!r}")


if __name__ == "__main__":
    main()
