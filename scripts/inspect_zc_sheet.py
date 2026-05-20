"""Inspecte la feuille ZC pour comprendre le tableau d'interpolation Taux ZC."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WB_PATH = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"


def main() -> None:
    wb_f = load_workbook(WB_PATH, data_only=False, read_only=False)
    wb_v = load_workbook(WB_PATH, data_only=True, read_only=False)
    for sn in ("ZC", "Courbe des taux"):
        if sn not in wb_f.sheetnames:
            continue
        ws_f = wb_f[sn]
        ws_v = wb_v[sn]
        print(f"\n=== {sn} dim={ws_f.max_row}x{ws_f.max_column} ===")
        for r in range(1, min(ws_f.max_row, 60) + 1):
            for c in range(1, min(ws_f.max_column, 15) + 1):
                vf = ws_f.cell(row=r, column=c).value
                vv = ws_v.cell(row=r, column=c).value
                if vf is None and vv is None:
                    continue
                col = ws_f.cell(row=r, column=c).coordinate
                print(f"  {col:<6} formule={vf!r:<60} valeur={vv!r}")


if __name__ == "__main__":
    main()
