"""Dump des blocs autour des codes 9424/9500 dans 'Ammortissable' du classeur de référence.

Affiche, pour chaque code :
- la cellule code et son label
- les lignes situées entre le code et le code suivant
- pour chaque ligne du bloc, la formule (data_only=False) et la valeur (data_only=True)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WB_PATH = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"


def col_letter(idx: int) -> str:
    out = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        out = chr(65 + rem) + out
    return out


def dump(ws_f, ws_v, anchor_row: int, max_rows: int = 30, max_cols: int = 40) -> None:
    """Dump des cellules non vides autour de l'ancre."""
    print(f"\n--- Bloc autour de la ligne {anchor_row} (lecture {max_rows} lignes x {max_cols} colonnes) ---")
    for r in range(anchor_row, anchor_row + max_rows):
        row_f = ws_f[r]
        row_v = ws_v[r]
        for c in range(1, max_cols + 1):
            cf = row_f[c - 1].value if c - 1 < len(row_f) else None
            cv = row_v[c - 1].value if c - 1 < len(row_v) else None
            if cf is None and cv is None:
                continue
            coord = f"{col_letter(c)}{r}"
            print(f"  {coord:<6} formule={cf!r:<60} valeur={cv!r}")


def main() -> None:
    wb_f = load_workbook(WB_PATH, data_only=False, read_only=False)
    wb_v = load_workbook(WB_PATH, data_only=True, read_only=False)
    ws_f = wb_f["Ammortissable"]
    ws_v = wb_v["Ammortissable"]
    for code, row in {"9424": 201, "9500": 1077}.items():
        print("\n" + "=" * 80)
        print(f"CODE {code} : ligne d'ancrage {row} (Ammortissable)")
        print("=" * 80)
        dump(ws_f, ws_v, row, max_rows=30, max_cols=60)


if __name__ == "__main__":
    main()
