#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compare le prix ATP Python avec le classeur de référence type **2026-PRICER_WG_CORRIGE**.

Usage (à la racine du projet ``pricer``) ::
    python scripts/extraire_prix_atp_pricer_wg.py
    python scripts/extraire_prix_atp_pricer_wg.py "2026-PRICER_WG_CORRIGE.xlsx" --code 200792

Placez le fichier ``.xlsx`` dans le dossier du projet (ou passez le chemin absolu).
Ouvrez une fois le classeur dans Excel et enregistrez-le pour que les **formules** aient des
**valeurs calculées** (``data_only``) ; sinon seules les formules texte seront visibles.

Formule de référence (ex. ligne 30) ::
    =prix_atp(AF$1; N; AB; O; M; K; Y; X; R; S; P; D; AA)
soit : liquidation (AF1), émission, jouissance, échéance, taux, nominal, prem_j, mode,
       périodicité coupon, périodicité cap, maturité, rendement, base.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

# Racine projet
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import pandas as pd

from pricing_atp import date_jouissance_wg_depuis_emission_echeance, prix_atp_dbt
from valuation_zc_obligations import (
    _arrondi_taux_decimal_excel,
    _arrondi_taux_facial_pct_wg,
    _parse_datetime_loose,
    _periodicite_coupon_depuis_valeur,
    _to_float_loose,
    detecter_colonnes_base_titre,
)


def _norm_sheet(name: str) -> str:
    return re.sub(r"\s+", "", str(name).lower().replace("é", "e").replace("è", "e"))


def _trouver_feuille(xl: pd.ExcelFile) -> str:
    want = "lignesdetenues"
    for s in xl.sheet_names:
        if _norm_sheet(s) == want or ("ligne" in _norm_sheet(s) and "detenu" in _norm_sheet(s)):
            return s
    raise ValueError(f"Aucune feuille type « Lignes_détenues » parmi : {xl.sheet_names}")


def _trouver_colonne_code(df: pd.DataFrame) -> str:
    for c in df.columns:
        if str(c).strip().upper().replace("É", "E") in ("CODE", "CODE MAROCLEAR", "CODE_MAROCLEAR"):
            return str(c)
    for c in df.columns:
        if "code" in str(c).lower() and "type" not in str(c).lower():
            return str(c)
    raise ValueError("Colonne CODE introuvable. En-têtes : " + ", ".join(map(str, df.columns[:25])))


def _as_date(v) -> date | None:
    d = _parse_datetime_loose(v)
    if d is None:
        return None
    return d.date() if hasattr(d, "date") else d


def _cellule_prix_probable(df: pd.DataFrame, row: pd.Series) -> str | None:
    """Nom de colonne contenant un prix / valorisation (valeur numérique proche du nominal)."""
    for c in df.columns:
        cl = str(c).lower()
        if any(
            x in cl
            for x in (
                "prix",
                "valorisation",
                "valeur",
                "clean",
                "cours",
                "quote",
            )
        ):
            if "dirty" in cl and "clean" not in cl:
                continue
            return str(c)
    return None


def _lire_ligne_wg(
    df: pd.DataFrame,
    code: str,
    col_code: str,
) -> tuple[pd.Series, int] | tuple[None, None]:
    code_s = str(code).strip()
    for i, row in df.iterrows():
        v = row[col_code]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip().replace(" ", "").replace("\xa0", "")
        if s.replace(".0", "") == code_s or s.startswith(code_s):
            try:
                if int(float(str(v).replace(",", "."))) == int(float(code_s)):
                    return row, int(i)
            except ValueError:
                if code_s in s:
                    return row, int(i)
    return None, None


def _valeur(row: pd.Series, noms: list[str], df: pd.DataFrame) -> object | None:
    cart = {str(c).strip().lower().replace("é", "e"): c for c in df.columns}
    for n in noms:
        k = n.lower().replace("é", "e")
        if k in cart:
            return row[cart[k]]
    for c in df.columns:
        cl = str(c).lower().replace("é", "e")
        for n in noms:
            if n.lower() in cl:
                return row[c]
    return None


def construire_appel_prix_atp(
    row: pd.DataFrame,
    df: pd.DataFrame,
    d_liq_global: date,
    cols_map: dict[str, Any] | None = None,
) -> dict:
    """
    Mappe les colonnes usuelles Maroclear / WG vers ``prix_atp_dbt``.
    Si ``cols_map`` est fourni (sortie de ``detecter_colonnes_base_titre``), on utilise les mêmes
    règles que la valorisation API (ex. **TAUX** avant **VALEUR TAUX**, rendement « nouvelle courbe »).
    """
    cm = cols_map or {}

    def cell(key: str) -> object | None:
        col = cm.get(key)
        if col and col in row.index:
            return row[col]
        return None

    def g(*aliases: str):
        return _valeur(row, list(aliases), df)

    d_em = _as_date(cell("col_date_emission") or g("DATE_EMISSION", "DATE EMIS", "DATE D'EMISSION", "EMISSION"))
    d_mat = _as_date(
        cell("col_date_echeance")
        or g("DATE_ECHEANCE", "DATE ECHEANCE", "DATE_ÉCHÉANCE", "ECHEANCE", "ÉCHÉANCE")
    )
    if d_mat is None:
        d_mat = _as_date(g("Date d'échéance", "Date d'echeance"))
    tc_col = cm.get("col_taux_coupon")
    tc = _to_float_loose(row[tc_col] if tc_col and tc_col in row.index else None)
    if tc is None:
        tc = _to_float_loose(g("VALEUR_TAUX", "TAUX", "TAUX FACIAL", "TAUX_COUPON"))
    nom = _to_float_loose(cell("col_nominal") or g("NOMINAL", "VN", "ENCOURS"))
    pj_raw = cell("col_premier_j") or g("PREMIER_J_INCLUS", "PREM_J_INCLUS", "PREMIER J")
    mode_raw = cell("col_mode_valo") or g("METHODE_VALO", "MODE_VALO", "MODE_VALORISATION", "MODE")
    per_raw = cell("cper") or g("PERIODE_COUPON", "PERIODICITE_COUPO", "PERIODICITE COUPON")
    pcap = cell("col_pcap") or g("PERIODICITE_REMBOURS", "PERIODICITE_CAP", "PERIODICITE CAP")
    mat_raw = cell("col_maturite_ct") or g("MATURITE_CT", "MATURITE", "MATURITE TITRE")
    rcol = cm.get("col_rendement_atp")
    rend = _to_float_loose(row[rcol] if rcol and rcol in row.index else None)
    if rend is None:
        rend = _to_float_loose(g("RENDEMENT_ACTUARIEL", "TAUX_ACTUARIEL", "RENDEMENT", "TAUX VALO"))
    base_raw = _to_float_loose(cell("col_base_atp") or g("BASE_ACTUARIEL", "BASE_ATP", "BASE"))

    if tc is not None and abs(tc) > 1.0:
        tc = float(tc) / 100.0
    if rend is not None and abs(rend) > 1.0:
        rend = float(rend) / 100.0
    if tc is not None and math.isfinite(float(tc)):
        tc = _arrondi_taux_facial_pct_wg(float(tc))
    if rend is not None and math.isfinite(float(rend)):
        rend = _arrondi_taux_decimal_excel(float(rend))

    pj = False
    if pj_raw is not None:
        pj = str(pj_raw).strip().upper().startswith("O") or pj_raw in (1, True, "1", "Y")

    cap_fin = True
    if pcap is not None and str(pcap).strip():
        v = str(pcap).upper()
        cap_fin = "FIN" in v or v.strip().startswith("F") or "INFINE" in v.replace(" ", "")

    per = _periodicite_coupon_depuis_valeur(per_raw) if per_raw is not None else 1

    mt_ct = None
    if mat_raw is not None:
        try:
            m = int(round(float(mat_raw)))
            if m in (13, 26, 52):
                mt_ct = m
        except (TypeError, ValueError):
            pass

    base_atp = 1
    if base_raw is not None and int(round(float(base_raw))) in (1, 2):
        base_atp = int(round(float(base_raw)))

    de = d_em or d_liq_global
    dj = date_jouissance_wg_depuis_emission_echeance(de, d_mat) if d_mat is not None else de

    return {
        "date_liquidation": d_liq_global,
        "date_emission": de,
        "date_jouissance": dj,
        "date_echeance": d_mat,
        "taux_coupon_annuel": float(tc) if tc is not None else float("nan"),
        "nominal": float(nom) if nom is not None else float("nan"),
        "premier_j_inclus": pj,
        "mode_valorisation": str(mode_raw or "A").strip(),
        "periodicite_cp": per,
        "periodicite_cap_fin": cap_fin,
        "rendement_annuel_effectif": float(rend) if rend is not None else float("nan"),
        "maturite_semaines_ct": mt_ct,
        "actuariel_base": base_atp,
    }


def _date_liquidation_af1(path: Path, sheet: str) -> date | None:
    """Lit ``AF1`` comme dans ``=prix_atp(AF$1;…)`` (feuille Lignes_détenues)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        return _as_date(ws["AF1"].value)
    except Exception:
        return None
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def main() -> int:
    ap = argparse.ArgumentParser(description="Compare prix ATP Excel vs Python (feuille Lignes_détenues).")
    ap.add_argument(
        "xlsx",
        nargs="?",
        default=None,
        help="Chemin du classeur (défaut : cherche *PRICER*WG*.xlsx à la racine du projet)",
    )
    ap.add_argument("--code", default="200792", help="CODE titre (ex. 200792)")
    args = ap.parse_args()

    racine = _ROOT
    path = Path(args.xlsx) if args.xlsx else None
    if path is None or not path.is_file():
        cands = (
            sorted(racine.glob("*PRICER*WG*.xlsm"))
            + sorted(racine.glob("*PRICER*WG*.xlsx"))
            + sorted(racine.glob("*pricer*wg*.xlsm"))
            + sorted(racine.glob("*pricer*wg*.xlsx"))
        )
        if not cands:
            cands = sorted(racine.glob("*.xlsm")) + sorted(racine.glob("*.xlsx"))
        if not cands:
            print(
                "Aucun .xlsx trouvé. Copiez **2026-PRICER_WG_CORRIGE.xlsx** (ou équivalent) "
                f"dans :\n  {racine}\n"
                "Puis relancez :\n  python scripts/extraire_prix_atp_pricer_wg.py votre_fichier.xlsx",
                file=sys.stderr,
            )
            return 2
        path = cands[0]
        print(f"Fichier utilisé : {path.name}\n")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = _trouver_feuille(xl)
    df = pd.read_excel(path, sheet_name=sheet, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    col_code = _trouver_colonne_code(df)
    row, idx = _lire_ligne_wg(df, args.code, col_code)
    if row is None:
        print(f"CODE {args.code!r} introuvable dans {path.name} / {sheet!r}.", file=sys.stderr)
        return 1

    # Date de valorisation : d’abord **AF1** (référence VBA), puis colonnes type date valo.
    d_liq = _date_liquidation_af1(path, sheet)
    if d_liq is None:
        for c in df.columns:
            cl = str(c).lower()
            if "valo" in cl and "date" in cl:
                v0 = df[c].iloc[0] if len(df) > 0 else None
                d_liq = _as_date(v0)
                if d_liq:
                    break
    if d_liq is None:
        for c in df.columns:
            if "date" in str(c).lower() and "echeance" not in str(c).lower():
                v0 = df[c].iloc[0]
                d_liq = _as_date(v0)
                if d_liq:
                    break
    if d_liq is None:
        print(
            "Date de liquidation introuvable (essayez **AF1** sur la feuille ou une date valorisation en ligne 1).",
            file=sys.stderr,
        )
        return 1

    cm = detecter_colonnes_base_titre(df)
    kw = construire_appel_prix_atp(row, df, d_liq, cols_map=cm)

    def _bad(x) -> bool:
        return x is None or (isinstance(x, float) and (math.isnan(x) or not math.isfinite(x)))

    if (
        kw.get("date_echeance") is None
        or _bad(kw.get("nominal"))
        or _bad(kw.get("taux_coupon_annuel"))
        or _bad(kw.get("rendement_annuel_effectif"))
    ):
        print("Paramètres incomplets (échéance, nominal, taux facial ou rendement manquant) :", kw, file=sys.stderr)
        print("Colonnes disponibles :", list(df.columns), file=sys.stderr)
        return 1

    # Colonne prix Excel (valeur affichée)
    col_prix = _cellule_prix_probable(df, row)
    prix_excel = None
    if col_prix and col_prix in row.index:
        prix_excel = _to_float_loose(row[col_prix])

    out_vba = prix_atp_dbt(**kw, taux_coupon_comme_vba=True)
    out_std = prix_atp_dbt(**kw, taux_coupon_comme_vba=False)

    print(f"Fichier      : {path}")
    print(f"Feuille      : {sheet}")
    print(f"Ligne (idx)  : {idx}  CODE={args.code}")
    print(f"Liquidation  : {d_liq}")
    print(f"Colonnes lues: échéance={kw.get('date_echeance')}  nominal={kw.get('nominal')}  "
          f"taux={kw.get('taux_coupon_annuel')}  R={kw.get('rendement_annuel_effectif')}  "
          f"base={kw.get('actuariel_base')}  coupon_VBA_style=True/False comparés")
    if col_prix:
        print(f"Prix Excel ({col_prix}): {prix_excel}")
    else:
        print("Prix Excel   : (colonne prix non détectée — ajoutez un en-tête « Prix » ou « Valorisation »)")
    print(f"Prix Py VBA  : {out_vba.get('prix_clean')}  (flux comme VBA nominal*taux)")
    print(f"Prix Py std  : {out_std.get('prix_clean')}  (facial annuel / périodicité)")
    if prix_excel is not None and out_vba.get("prix_clean"):
        ev, pv = float(prix_excel), float(out_vba["prix_clean"])
        print(f"Écart VBA    : {pv - ev:+.6f}")
    if prix_excel is not None and out_std.get("prix_clean"):
        ev, pv = float(prix_excel), float(out_std["prix_clean"])
        print(f"Écart std    : {pv - ev:+.6f}")

    print("\n--- Flux (mode VBA coupons) ---")
    print("Dates :", out_vba.get("flux_dates"))
    print("Montants:", out_vba.get("flux_montants"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
