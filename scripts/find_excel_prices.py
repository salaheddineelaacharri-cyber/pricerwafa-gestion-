"""Trouve les prix de référence 9424, 9500, 9351, 101006, 9651 dans le classeur."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

wb = load_workbook(wb_path, data_only=True, keep_vba=True)
TARGET_CODES = ("9424", "9500", "9351", "101006", "9651")

for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            sv = str(v).strip()
            for code in TARGET_CODES:
                if sv == code or sv == code + ".0":
                    # Print cell location, then list neighbour values along the row to find price.
                    coord = cell.coordinate
                    row_vals = []
                    for c in ws.iter_cols(min_row=cell.row, max_row=cell.row, min_col=1, max_col=min(40, ws.max_column)):
                        for cc in c:
                            row_vals.append((cc.coordinate, cc.value))
                    print(f"\n[{sn}] {coord} = {sv}")
                    for cc, vv in row_vals:
                        if vv is not None:
                            print(f"    {cc}: {vv!r}")
                    break
