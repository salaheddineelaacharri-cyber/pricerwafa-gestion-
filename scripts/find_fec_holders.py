"""Trouve quel(s) code(s) de Ammortissable correspondent à OBL FEC 06/12/2017 3.93% 15 ANS."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

wb_v = load_workbook(wb_path, data_only=True, keep_vba=True)
ws_v = wb_v["Ammortissable"]

# Print all rows that have "FEC 06/12/2017" or rate 3.93%
for r in range(1, ws_v.max_row + 1):
    vals = []
    for c in range(1, 7):
        v = ws_v.cell(r, c).value
        if v is not None:
            vals.append((chr(64+c), v))
    if not vals:
        continue
    text = " ".join(str(v) for _, v in vals)
    if "FEC" in text and ("3.93" in text or "0.0393" in text or "06/12" in text or "2017" in text):
        print(f"  Row {r}: {vals}")

# Also list all "Prix" rows
print("\n--- All Prix rows ---")
for r in range(1, ws_v.max_row + 1):
    if ws_v.cell(r, 2).value == "Prix":
        c = ws_v.cell(r, 3).value
        # Get the code from a row above
        # Format: bond header is typically 14 rows above Prix
        # Code is on row r-14, col C usually
        code_above = ws_v.cell(r - 14, 3).value
        print(f"  Row {r}: code={code_above!r}  Prix={c!r}")
