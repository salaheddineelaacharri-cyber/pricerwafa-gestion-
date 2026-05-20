"""Recherche 9351 et 47586 sur toutes les feuilles pour comprendre quel onglet le calcule."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

wb_v = load_workbook(wb_path, data_only=True, keep_vba=True)
wb_f = load_workbook(wb_path, data_only=False, keep_vba=True)

for sn in wb_v.sheetnames:
    ws = wb_v[sn]
    found_codes = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            sv = str(v)
            if sv == "9351" or sv == "9351.0":
                found_codes.append((cell.row, cell.column, "9351"))
            elif "47586" in sv and "." in sv:
                found_codes.append((cell.row, cell.column, sv))
    if found_codes:
        print(f"\n[{sn}] {len(found_codes)} hits:")
        for r, c, val in found_codes[:30]:
            print(f"  Row {r} Col {c}: {val}")
