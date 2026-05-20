"""Compare le moteur backend à toutes les références de CONTROLE_PRICER_MANAR.

Affiche un récap (n_match / n_total) et liste les codes avec écart > 0.01."""

from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from backend import main as api  # noqa: E402

wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"
wb = load_workbook(wb_path, data_only=True, keep_vba=True)
ws = wb["CONTROLE_PRICER_MANAR"]

# Lecture: col A = code, col D = prix réf 4 décimales (D86=100577.6813)
refs: dict[str, float] = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None:
        continue
    code_v = row[0]
    ref_v = row[3] if len(row) > 3 else None
    try:
        code_s = str(code_v).strip()
        if code_s.endswith(".0"):
            code_s = code_s[:-2]
        if code_s and ref_v is not None:
            refs[code_s] = float(ref_v)
    except (TypeError, ValueError):
        continue
print(f"Loaded {len(refs)} reference prices from CONTROLE_PRICER_MANAR")

pillars = api._extraire_piliers_depuis_histo(ROOT, "2026-03-26", "MAR_JJ")
curve = api.CurveRequest(
    short=pillars["short"],
    long=pillars["long"],
    joint_days=325,
    max_days=11000,
    step_short=50,
    step_long=100,
)
req = api.MarcheValorizeRequest(
    valuation_date="2026-03-26",
    curve=curve,
    feuil1_pricer_tous=True,
)
res = api.marche_valorize(req)
if hasattr(res, "body"):
    res = json.loads(res.body)

rows = res.get("rows", []) or res.get("lignes", [])

def _norm(c) -> str:
    s = str(c or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

by_code: dict[str, dict] = {}
for r in rows:
    c = _norm(r.get("CODE") or r.get("code") or r.get("titre"))
    if c:
        by_code[c] = r

n_match = 0
n_diff = 0
n_missing = 0
diffs: list[tuple[str, float, float, float]] = []
for code, ref in refs.items():
    r = by_code.get(code)
    if not r:
        n_missing += 1
        continue
    pa = r.get("Prix arrondi") or r.get("prix_arrondi") or r.get("Prix clean")
    if pa is None:
        n_missing += 1
        continue
    try:
        pa_f = float(pa)
    except (TypeError, ValueError):
        n_missing += 1
        continue
    delta = pa_f - float(ref)
    if abs(delta) <= 0.01:
        n_match += 1
    else:
        n_diff += 1
        diffs.append((code, pa_f, float(ref), delta))

print(f"\nMatches (|delta| <= 0.01): {n_match}")
print(f"Diffs   (|delta|  > 0.01): {n_diff}")
print(f"Missing in backend output: {n_missing}")

if diffs:
    diffs.sort(key=lambda t: -abs(t[3]))
    print("\nTop divergences (|delta| desc, full list):")
    print(f"  {'code':<10} {'moteur':>16} {'classeur':>16} {'ecart':>12}")
    for code, m, ref, d in diffs:
        print(f"  {code:<10} {m:>16.4f} {ref:>16.4f} {d:>12.4f}")
