"""Trouve le titre FEC 06/12/2017 (qui correspond à 9351) dans Ammortissable."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

wb_v = load_workbook(wb_path, data_only=True, keep_vba=True)
wb_f = load_workbook(wb_path, data_only=False, keep_vba=True)
ws_v = wb_v["Ammortissable"]
ws_f = wb_f["Ammortissable"]

# Search for FEC, EFEC, 9351, 9149 etc. in Ammortissable (columns A-E typically)
for row in ws_v.iter_rows(values_only=False):
    for cell in row:
        v = cell.value
        if v is None:
            continue
        s = str(v)
        if "FEC" in s.upper() or "06/12/2017" in s:
            r = cell.row
            print(f"\n[Row {r}] {cell.coordinate} = {v!r}")
            # Print neighbour
            for c in range(1, 8):
                vv = ws_v.cell(r, c).value
                if vv is not None:
                    print(f"  {chr(64+c)}{r} = {vv!r}")
            break
