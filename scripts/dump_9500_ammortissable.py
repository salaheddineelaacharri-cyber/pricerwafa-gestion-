"""Dump le bloc Ammortissable pour 9500 et 9351 — formules + valeurs calculées."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
wb_path = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"

wb_v = load_workbook(wb_path, data_only=True, keep_vba=True)
wb_f = load_workbook(wb_path, data_only=False, keep_vba=True)

ws_v = wb_v["Ammortissable"]
ws_f = wb_f["Ammortissable"]

TARGETS = ("9500", "9351", "9424", "101006")

# Find the row containing each target code (typically column C).
for code in TARGETS:
    print(f"\n========== Code {code} ==========")
    found_rows = []
    for row in ws_v.iter_rows(min_col=1, max_col=5, values_only=False):
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() == code:
                found_rows.append(cell.row)
    if not found_rows:
        # search all columns
        for row in ws_v.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if v is not None and str(v).strip() == code:
                    found_rows.append(cell.row)
                    break
    if not found_rows:
        print("    (introuvable)")
        continue
    r0 = min(found_rows)
    # The block usually spans ~14 rows (header + Date, Capital, Intérêts, Flux, Flux restant, durée, Taux ZC, Prime, Taux d'actu, Flux actualisé)
    print(f"  Row code = {r0}")
    last_col = min(ws_v.max_column, 60)
    for r in range(r0 - 1, r0 + 16):
        if r < 1:
            continue
        row_v = [ws_v.cell(r, c).value for c in range(1, last_col + 1)]
        row_f = [ws_f.cell(r, c).value for c in range(1, last_col + 1)]
        # only print if first column has a label
        first = row_v[0] if row_v[0] is not None else (row_v[1] if len(row_v) > 1 else None)
        line = []
        for cc, (vv, ff) in enumerate(zip(row_v, row_f)):
            if vv is None and ff is None:
                continue
            line.append(f"{chr(ord('A')+cc) if cc<26 else 'A'+chr(ord('A')+cc-26)}{r}={vv!r}")
        print("    ".join(line) if line else "")
