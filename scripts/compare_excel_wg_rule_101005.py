"""Compare WG Ammortissable (formules Exactes) vs grille Python pour 101005 × dates."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from backend import main as api
from obligation_amort_schedule import _round_excel, construire_tables_amortissement_pour_valorisation
from valuation_zc_obligations import charger_courbe_zc_depuis_fichier, valoriser_dataframe_base_titre


def wg_excel_pv_breakdown() -> None:
    WB = ROOT / "2026-PRICER_WG_CORRIGE .xlsm"
    wb_v = load_workbook(WB, data_only=True, keep_vba=True)
    wb_f = load_workbook(WB, data_only=False, keep_vba=True)
    ws_v = wb_v["Ammortissable"]
    ws_f = wb_f["Ammortissable"]
    print("C1 valo (valeur):", ws_v["C1"].value, "| formule:", ws_f["C1"].value)
    print("D1025 formule:", ws_f["D1025"].value)
    print("D1026 formule:", getattr(ws_f["D1026"].value, "text", ws_f["D1026"].value))
    print("D1028 formule:", ws_f["D1028"].value)
    print("D1029 formule:", ws_f["D1029"].value)
    print("C1031 formule:", ws_f["C1031"].value)

    cols = list("DEFGHIJ")
    rows: dict = {}
    for col in cols:
        rows[col] = {
            "flux": float(ws_v[f"{col}1024"].value or 0.0),
            "duree": float(ws_v[f"{col}1025"].value or 0.0),
            "zc_dec": float(ws_v[f"{col}1026"].value or 0.0),
            "prime_dec": float(ws_v[f"{col}1027"].value or 0.0),
            "actu_dec": float(ws_v[f"{col}1028"].value or 0.0),
            "pv_4": float(ws_v[f"{col}1029"].value or 0.0),
        }
        if rows[col]["flux"] > 0:
            raw_pv = rows[col]["flux"] / (
                (1.0 + rows[col]["actu_dec"]) ** rows[col]["duree"]
            )
            rows[col]["round4_check"] = _round_excel(raw_pv, 4)
    prix = float(ws_v["C1031"].value)
    s_pv = sum(r["pv_4"] for r in rows.values())
    wb_v.close()
    wb_f.close()
    print("\n=== WG fichier (bloc 101005) — regle Prix ===")
    print("Prix C1031 (= SOMME des Flux act. arrondis 4 dec.):", prix)
    print("Somme cellules D1029:J1029:", s_pv)
    for col, r in rows.items():
        if r["flux"] <= 0:
            continue
        print(
            col,
            f"flux={r['flux']:.6f}",
            f"d={r['duree']:.12g}",
            f"actu={r['actu_dec']:.8f}",
            f"pv_cell={r['pv_4']:.4f}",
            f"recalc_round4={r['round4_check']:.4f}",
        )


def python_table_for_date(iso: str) -> None:
    zc_path = ROOT / "pricing/curves/courbe_zc.py"
    courbe = charger_courbe_zc_depuis_fichier(zc_path)
    pillars = api._extraire_piliers_depuis_histo(ROOT, iso, "MAR_JJ")
    req_curve = api.CurveRequest(
        short=[api.PillarShort(**p) for p in pillars["short"]],
        long=[api.PillarLong(**p) for p in pillars["long"]],
        joint_days=float(pillars.get("joint_days", 325.0)),
        max_days=11000,
        step_short=50,
        step_long=100,
    )
    curve_tracee = api._make_curve(req_curve)
    schedule_zc = api._schedule_table_records(curve_tracee, root=ROOT, date_courbe=iso)
    fn_j = lambda j, rows=schedule_zc: api._interp_taux_zc_actuariel_depuis_schedule_jours(float(j), rows)  # noqa: E731
    fn_a = lambda a, rows=schedule_zc: api._interp_taux_zc_depuis_schedule_annuel(float(a), rows)  # noqa: E731

    bam_cc, bam_cl = api._courbes_bam_depuis_requete(req_curve)

    def ts_amort(j: float) -> float:
        return float(
            api.taux_secondaire_interpole_formule_b(
                float(j),
                bam_cc,
                bam_cl,
                ndigits=api.NDIGITS_TAUX_SECONDAIRE_FORMULE_B_DEFAUT,
            )
        )

    xlsx = api.resoudre_fichier_base_titre_oblig(ROOT, None)
    df_in = api._charger_base_titre_oblg_cache(xlsx, ["101005"])
    if df_in.empty:
        print(f"{iso}: base titre vide pour 101005")
        return

    df_out, det_slice = valoriser_dataframe_base_titre(
        df_in,
        courbe,
        valuation_date=iso,
        bam_courbe_court=bam_cc,
        bam_courbe_long=bam_cl,
        progress_label=None,
    )
    raw_l = api._df_to_records(df_out)
    rows_ui = [api._row_to_marche_ui(r, iso) for r in raw_l]
    col_code = "CODE"
    amort_tables = construire_tables_amortissement_pour_valorisation(
        xlsx,
        raw_l,
        rows_ui,
        valuation_date=iso,
        taux_secondaire_a_j=ts_amort,
        taux_zc_schedule_j=fn_j,
        taux_zc_schedule_a=fn_a,
        df_work=df_out,
        col_code_fichier=col_code,
        det_cols=det_slice,
        codes_filter=["101005"],
    )
    tab = amort_tables[0] if amort_tables else {}
    rows_by_lbl = {str(r.get("label")): r for r in (tab.get("rows") or []) if isinstance(r, dict)}

    flux_act = rows_by_lbl.get("Flux actualisé", {}).get("values") or []
    ta = rows_by_lbl.get("Taux d'actualisation", {}).get("values") or []
    fr = rows_by_lbl.get("Flux restant", {}).get("values") or []
    du = rows_by_lbl.get("durée", {}).get("values") or []
    zc = rows_by_lbl.get("Taux ZC", {}).get("values") or []

    sum_round4 = sum(float(x) for x in flux_act if x not in (None, ""))
    prix_sum6 = tab.get("prix_somme_flux_actualises")

    print(f"\n=== Python (même pipeline que marché) {iso} ===")
    print("prix_somme_flux_actualises (somme PV pleine precision, hors arrondi colonne):", prix_sum6)
    print("Somme cellules Flux actualise (arrondi 4 dec. affichage):", round(sum_round4, 6))
    print("ecart somme_affichage - somme_HP:", round(sum_round4 - float(prix_sum6 or 0.0), 8))

    # Quelques colonnes non nulles
    def _nz(vals: list) -> list[tuple[int, float]]:
        out: list[tuple[int, float]] = []
        for i, v in enumerate(vals):
            try:
                f = float(v)
            except (TypeError, ValueError):
                continue
            if abs(f) > 1e-9:
                out.append((i, f))
        return out[:12]

    print("durée (non nul):", _nz(list(du)))
    print("Taux ZC % (non nul):", _nz(list(zc)))
    print("Taux actu % (non nul):", _nz(list(ta)))
    print("Flux restant (non nul):", _nz(list(fr)))
    print("Flux act. (non nul):", _nz(list(flux_act)))


def main() -> None:
    wg_excel_pv_breakdown()
    for iso in ("2026-03-26", "2026-03-06", "2026-01-02"):
        python_table_for_date(iso)


if __name__ == "__main__":
    main()
